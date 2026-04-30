"""
Digzio Property Bulk Import - Robust Version with SSL Retry
Handles SSL connection drops and rate limits gracefully.
"""
import openpyxl, time, json, sys
from datetime import datetime
import urllib3
urllib3.disable_warnings()

# Use requests with retry adapter
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

SPREADSHEET = "/home/ubuntu/upload/MasterStudentHousingDatasetUpdatedDFIVersion.xlsx"
BASE_URL = "https://www.digzio.co.za"
AUTH_URL = f"{BASE_URL}/api/v1/auth"
PROP_URL = f"{BASE_URL}/api/v1/properties"
DEFAULT_PASSWORD = "Provider@2026!"

PROVIDERS = [
    {"name": "Cape Student Residences",          "email": "info@capestudentres.co.za"},
    {"name": "Roomza Demo Provider",             "email": "provider@roomza.co.za"},
    {"name": "Gauteng Student Housing",          "email": "admin@gautenghousing.co.za"},
    {"name": "Stellenbosch Accommodation Group", "email": "bookings@stellaccgroup.co.za"},
]

PROPERTY_TYPE_MAP = {
    "apartment": "APARTMENT", "flat": "APARTMENT", "studio": "APARTMENT",
    "house": "HOUSE", "townhouse": "HOUSE", "cluster": "HOUSE",
    "room": "ROOM", "boarding": "ROOM", "backroom": "ROOM",
    "student residence": "STUDENT_RESIDENCE", "residence": "STUDENT_RESIDENCE",
    "hostel": "HOSTEL",
}

def make_session():
    """Create a requests session with retry logic"""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=2,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET", "POST", "PATCH"],
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=4, pool_maxsize=4)
    session.mount("https://", adapter)
    session.verify = False
    return session

def normalise_type(raw):
    if not raw:
        return "APARTMENT"
    raw_lower = str(raw).lower()
    for k, v in PROPERTY_TYPE_MAP.items():
        if k in raw_lower:
            return v
    return "APARTMENT"

def generate_description(row):
    name = str(row.get("Property Name") or "This property").strip()
    suburb = str(row.get("Suburb") or "").strip()
    city = str(row.get("City") or "").strip()
    ptype = str(row.get("Property Type") or "accommodation").lower()
    price = row.get("Min Rent (R)")
    beds = row.get("Unit Count")
    nsfas = str(row.get("NSFAS Accredited") or "").strip().lower() in ["yes", "true", "1"]
    nsfas_str = "NSFAS-accredited and " if nsfas else ""
    location = f"{suburb}, {city}".strip(", ")
    desc = (f"{name} is a {nsfas_str}student accommodation facility located in {location}. "
            f"Offering quality {ptype} units at competitive rates")
    if price:
        desc += f" from R{price}/month"
    if beds:
        desc += f", with {beds} units available"
    desc += ". Ideal for students seeking safe, comfortable, and affordable housing close to campus."
    return desc

def login(session, email):
    for attempt in range(3):
        try:
            resp = session.post(f"{AUTH_URL}/login",
                                json={"email": email, "password": DEFAULT_PASSWORD},
                                timeout=20)
            if resp.status_code == 200:
                data = resp.json()
                token = data.get("token") or data.get("data", {}).get("token")
                if token:
                    print(f"  LOGIN OK: {email}", flush=True)
                    return token
            print(f"  LOGIN FAILED {resp.status_code}: {email}", flush=True)
            return None
        except Exception as e:
            if attempt < 2:
                print(f"  LOGIN RETRY {attempt+1}: {email} - {e}", flush=True)
                time.sleep(5)
            else:
                print(f"  LOGIN ERROR: {email} - {e}", flush=True)
    return None

def load_properties():
    wb = openpyxl.load_workbook(SPREADSHEET, read_only=True, data_only=True)
    ws = wb["MASTER_DATASET"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    props = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        lat = r.get("Latitude")
        lon = r.get("Longitude")
        if not lat or not lon:
            skipped += 1
            continue
        props.append(r)
    print(f"Loaded {len(props)} properties (skipped {skipped} without GPS)", flush=True)
    return props

def create_property(session, token, row):
    lat = float(row.get("Latitude", 0))
    lon = float(row.get("Longitude", 0))
    price = row.get("Min Rent (R)") or 3000
    beds = row.get("Unit Count") or 1
    nsfas = str(row.get("NSFAS Accredited") or "").strip().lower() in ["yes", "true", "1"]
    address = str(row.get("Address") or "").strip()
    suburb = str(row.get("Suburb") or "").strip()
    city = str(row.get("City") or "").strip()
    address_line_1 = address if address else (f"{suburb}, {city}".strip(", ") or "Address not specified")
    
    payload = {
        "title": str(row.get("Property Name") or "Student Accommodation").strip(),
        "description": generate_description(row),
        "address_line_1": address_line_1,
        "city": city,
        "province": str(row.get("Province") or "").strip(),
        "postal_code": str(row.get("Postal Code") or "0000").strip() or "0000",
        "latitude": lat,
        "longitude": lon,
        "property_type": normalise_type(row.get("Property Type")),
        "total_beds": int(float(str(beds))) if beds else 1,
        "available_beds": int(float(str(beds))) if beds else 1,
        "base_price_monthly": float(str(price)) if price else 3000,
        "is_nsfas_accredited": nsfas,
        "amenities": ["wifi", "security"],
    }
    
    hdrs = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    
    for attempt in range(3):
        try:
            resp = session.post(PROP_URL, json=payload, headers=hdrs, timeout=25)
            return resp.status_code, resp.json() if resp.status_code in [201, 409, 429] else {}
        except Exception as e:
            if attempt < 2:
                print(f"  CONN RETRY {attempt+1}: {row.get('Property Name', '?')} - {type(e).__name__}", flush=True)
                time.sleep(10)
                # Recreate session on SSL errors
                session = make_session()
            else:
                return 0, {"error": str(e)}
    return 0, {}

def main():
    print("=== DIGZIO PROPERTY BULK IMPORT (ROBUST) ===", flush=True)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n", flush=True)
    
    all_props = load_properties()
    
    # Login all providers
    session = make_session()
    tokens = []
    for p in PROVIDERS:
        token = login(session, p["email"])
        if token:
            tokens.append({"email": p["email"], "token": token, "count": 0})
    
    if not tokens:
        print("ERROR: No providers logged in. Aborting.", flush=True)
        return
    
    print(f"\n{len(tokens)} provider tokens ready. Starting import of {len(all_props)} properties...", flush=True)
    print("Rate: 1 request per 4.8s to stay under 200/15min window\n", flush=True)
    
    inserted = 0
    skipped_dup = 0
    errors = 0
    
    for i, prop in enumerate(all_props):
        # Rotate tokens round-robin
        tok = tokens[i % len(tokens)]
        
        status, data = create_property(session, tok["token"], prop)
        
        if status == 201:
            inserted += 1
            tok["count"] += 1
            if inserted % 25 == 0:
                print(f"  Progress: {inserted}/{len(all_props)} inserted ({inserted/((i+1)*4.8/60):.1f}/min), {errors} errors", flush=True)
        elif status == 409:
            skipped_dup += 1
        elif status == 429:
            print(f"  Rate limited at {i+1}. Waiting 65s...", flush=True)
            time.sleep(65)
            # Retry
            status, data = create_property(session, tok["token"], prop)
            if status == 201:
                inserted += 1
            elif status == 409:
                skipped_dup += 1
            else:
                errors += 1
                if errors <= 20:
                    print(f"  RETRY FAILED [{status}]: {prop.get('Property Name', 'unknown')}", flush=True)
        elif status == 0:
            errors += 1
            if errors <= 20:
                print(f"  CONN ERROR: {prop.get('Property Name', 'unknown')} - {data.get('error', '')[:80]}", flush=True)
            # Recreate session on connection errors
            session = make_session()
        else:
            errors += 1
            if errors <= 20:
                print(f"  ERROR [{status}]: {prop.get('Property Name', 'unknown')}", flush=True)
        
        time.sleep(4.8)
    
    print(f"\n=== IMPORT COMPLETE ===", flush=True)
    print(f"  Inserted:   {inserted}", flush=True)
    print(f"  Duplicates: {skipped_dup}", flush=True)
    print(f"  Errors:     {errors}", flush=True)
    print(f"  Total:      {inserted + skipped_dup + errors}", flush=True)
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)

if __name__ == "__main__":
    main()
